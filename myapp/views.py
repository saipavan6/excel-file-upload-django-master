from django.shortcuts import render
import openpyxl
from . import dal
import os




def CheckExceldata(request):
    if "GET" == request.method:
        return render(request, 'myapp/ExelReadOnly.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["GRADE X"]
        row_data2 = list()
        for i,row in enumerate(worksheet.iter_rows()):
            print(row)
            # if i == 0:
            #     continue
            t = tuple()
            for cell in row:
                t = t + (cell.value,)
            row_data2.append(t)

        return render(request, 'myapp/ExelReadOnly.html', {"excel_data": row_data2})




def index(request):
    if "GET" == request.method:
        return render(request, 'myapp/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        #sheets = wb.sheetnames
        #print(sheets)

        # getting a particular sheet
        worksheet = wb["GRADE X"]

        # getting active sheet
        #active_sheet = wb.active
        #print(active_sheet)

        # reading a cell
        #print(worksheet["A1"].value)

        #excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        row_data2 = list()
        for i,row in enumerate(worksheet.iter_rows()):

            if i == 0:
                continue
            #row_data = list()
            t = tuple()
            for cell in row:
                t = t + (cell.value,)

            #for changing file names
            # if row[2].value == None:
            #     break
            # if int(row[4].value) == 1:
            #     file_path = rf"C:\Users\saipa\Downloads\birla school photos\{str(row[3].value)}\Boys\_MG_{str(row[12].value)}.jpg"
            #     new_name = rf"C:\Users\saipa\Downloads\birla school photos\{str(row[3].value)}\Boys\{str(row[0].value)}.jpg"
            # elif int(row[4].value) == 2:
            #     file_path = rf"C:\Users\saipa\Downloads\birla school photos\{str(row[3].value)}\Girls\_MG_{str(row[12].value)}.jpg"
            #     new_name = rf"C:\Users\saipa\Downloads\birla school photos\{str(row[3].value)}\Girls\{str(row[0].value)}.jpg"
            #
            # #Call the function to rename the file if it exists
            # if row[12].value != None:
            #     data = rename_existing_file(file_path, new_name)
            #     t = t +  (data,)

            row_data2.append(t)
        data = dal.dbStudentList(row_data2)
        return render(request, 'myapp/index.html', {"excel_data": row_data2})









        #print(row_data2)

        # import pyodbc
        # server = 'SUVARNAMUKHI\SQLEXPRESS'
        # database_name = 'CampusMgmtNew'
        # user_name = 'sa'
        # password = 'sadguru'
        # conn = pyodbc.connect(
        #             'DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + server + ';DATABASE=' + database_name + ';UID=' + user_name + ';PWD=' + password)
        #
        # cursor = conn.cursor()
        # for i in row_data2:
        #     print(i)
        #     if i[0] == None:
        #         AdmissionNo = 'null'
        #     else:
        #         AdmissionNo = f"'{i[0]}'"
        #
        #     if i[1] == None:
        #         HallTicketNo = 'null'
        #     else:
        #         HallTicketNo = f"'{i[1]}'"
        #
        #     if i[5] == None:
        #         dob = 'null'
        #     else:
        #         dob = f"'{i[5]}'"
        #     if i[6] == None:
        #         FatherName = 'null'
        #     else:
        #         FatherName = f"'{i[6]}'"
        #
        #     if i[7] == None:
        #         MobileNo = 'null'
        #     else:
        #         MobileNo = f"'{i[7]}'"
        #     if i[8] == None:
        #         MotherName = 'null'
        #     else:
        #         MotherName = f"'{i[8]}'"
        #
        #     if i[9] == None:
        #         Email = 'null'
        #     else:
        #         Email = f"'{i[9]}'"
        #
        #     if i[2] == None:
        #         conn.commit()
        #         print("Data Successfully Inserted")
        #         conn.close()
        #         return render(request, 'myapp/index.html', {"excel_data":row_data2})
        #
        #     SQLCommand = f"""DECLARE @Regid int,@SuperId int,@BranchId int,@Academicid int,@Sectionsid int,@Standardsid int
        #     SET @SuperId = 20019
        #     select @BranchId = id from [CONFIG].[Branches] where SuperId = @SuperId and Orgtype =1
        #     select @Academicid = id from [CONFIG].[AcademicYears] where SuperId = @SuperId and CurrentActive = 1 and IsActive = 1
        #     select @Sectionsid = id from [CONFIG].[Sections] where SuperId = @SuperId and BranchId = @BranchId
        #     select @Standardsid = s.id from [CONFIG].[Programs] p
        #         inner join [CONFIG].[Courses] c on c.Programid = P.ID
        #         INNER JOIN [CONFIG].[Standards] S on s.CourseId = c.Id
        #         where p.SuperId = 20019 and s.Name = '{i[3]}' and
        #         p.BranchId = @BranchId
        #
        #     insert into [Campus].Registrations(Superid,AdmissionNo,HallTicketNo,FirstName,Gender,Dob,BranchId)
        #     values(@SuperId,{AdmissionNo},{HallTicketNo},'{i[2]}',{i[4]},{dob},@BranchId)
        #
        #     SET @Regid = SCOPE_IDENTITY()
        #
        #     insert into [Campus].[StudentTransactions](Academicid,Regid,Standardid,sectionid,ispromoted)
        #     values(@Academicid,@Regid,@Standardsid,@Sectionsid,0)
        #
        #     insert into [Campus].[GuardianDetails](Regid,FatherName,MobileNo,MotherName,Email )
        #     values(@Regid,{FatherName},{MobileNo},{MotherName},{Email})"""
        #
        #     # cursor = conn.cursor()
        #     # quary_execution = cursor.execute(SQLCommand)
        #
        #     # Processing Query
        #     cursor.execute(SQLCommand)
        #
        # conn.commit()
        # print("Data Successfully Inserted")
        # conn.close()
        #
        #
        # return render(request, 'myapp/index.html', {"excel_data":row_data2})


def rename_existing_file(file_path, new_name):
    if os.path.exists(file_path):
        file_name, file_extension = os.path.splitext(new_name)
        i = 1
        while os.path.exists(new_name):
            new_name = f"{file_name}_{i}{file_extension}"
            i += 1
        os.rename(file_path, new_name)
        return f"File renamed to {new_name}"
    else:
        return "File does not exist."
def CheckPhotoExistorNot(request):
    if "GET" == request.method:
        return render(request, 'myapp/checkphotoexistornot.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["7th"]
        row_data2 = list()
        for i, row in enumerate(worksheet.iter_rows()):
            if i == 0:
                continue
            t = tuple()
            for cell in row:
                t = t + (cell.value,)

            if int(row[4].value) == 1:
                file_path = rf"C:\Users\saipa\Desktop\birla school photos\{str(row[3].value)}\Boys\_MG_{str(row[12].value)}.jpg"
                new_name = rf"C:\Users\saipa\Desktop\birla school photos\{str(row[3].value)}\Boys\{str(row[0].value)}.jpg"
            elif int(row[4].value) == 2:
                file_path = rf"C:\Users\saipa\Desktop\birla school photos\{str(row[3].value)}\Girls\_MG_{str(row[12].value)}.jpg"
                new_name = rf"C:\Users\saipa\Desktop\birla school photos\{str(row[3].value)}\Girls\{str(row[0].value)}.jpg"

            #Call the function to rename the file if it exists
            if row[12].value != None:
                data = rename_existing_file(file_path, new_name)
                t = t +  (data,)
            row_data2.append(t)

        print(row_data2)
        return render(request, 'myapp/checkphotoexistornot.html', {"excel_data": row_data2})






